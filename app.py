import streamlit as st
import google.generativeai as genai
import gspread
from google_auth_oauthlib.flow import InstalledAppFlow
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from google.auth import exceptions as google_exceptions
from googleapiclient.discovery import build
from num2words import num2words
from datetime import datetime, timedelta, date
import json
import time
import os
import io 
import pandas as pd
import zipfile
from PyPDF2 import PdfReader, PdfWriter
from db_client import save_admission_record, get_funcionarios_financeiro, _to_decimal, save_payroll_record
from decimal import Decimal, ROUND_HALF_UP

def calcular_dias_uteis_proporcionais(adm_str, fech_date):
    """Calcula dias úteis (Seg-Sex) desde a admissão até o fechamento se no mesmo mês."""
    try:
        dt_adm = datetime.strptime(adm_str, '%d/%m/%Y').date()
        # Se admitido no mesmo mês/ano do fechamento
        if dt_adm.year == fech_date.year and dt_adm.month == fech_date.month:
            d_uteis = 0
            curr = dt_adm
            while curr <= fech_date:
                if curr.weekday() < 5: d_uteis += 1
                curr += timedelta(days=1)
            return d_uteis
    except: pass
    return 22



# ==========================================
# 1. CONFIGURAÇÕES INICIAIS E VARIÁVEIS DE SESSÃO
# ==========================================
st.set_page_config(page_title="Vicelos System", page_icon="🏗️", layout="wide")

if 'dados_extraidos' not in st.session_state: st.session_state.dados_extraidos = None
if 'funcs' not in st.session_state: st.session_state.funcs = []
if 'ultimo_kit' not in st.session_state: st.session_state.ultimo_kit = None
if 'ultimo_holerite' not in st.session_state: st.session_state.ultimo_holerite = None
if 'df_lote' not in st.session_state: st.session_state.df_lote = None

API_KEY_GEMINI = os.getenv("GEMINI_API_KEY", "")

API_KEY_GEMINI = os.getenv("GEMINI_API_KEY", "")

ID_PASTA_RAIZ = "1_w4HGrBnylar-vkiQTDT6ozW8KiGITZs" 
ID_PLANILHA = "1-VH1zGyTeEfJnvBhnq6ZlkyGF-G--FKXTwGHfrCvfRE" 
ID_MODELO_HOLERITE = "1PwSXH2NOxxPer4MOchV9zfpeyyemIrULas7J8c1jTrk"

MODELOS_ADMISSAO = {
    "01. Contrato de Trabalho": "1877g_glh9TZ5DFRrUpVPQD_R-s099tFVCljREUlPZs0",
    "02. Termo LGPD": "155FUL0maCZftqZ-T5O3xwoitBCO4-OZL_-0_cI9KDMk",
    "03. Vale Transporte": "1mUop1DJef-V15F-6Z-9ACQqEsT7qU2_iCm9V8dz2VLA",
    "04. Ficha de EPI": "18U92SoGwqDKalMRnkVay-pfo0QVOG-1ZF0nHiaRcFhk",
    "05. Acordo Compensação": "1p0nmiiZq7O22Clwq0dixYTUlTykBQTL73-LtF4LtKd0",
    "06. Acordo Prorrogação": "1XuOQ1Z9MIeotWrbYnsH7XoSCzc24_FJ6MtYt9f3L0eA"
}

FAIXAS_INSS = [
    (Decimal("1518.00"), Decimal("0.075")), 
    (Decimal("2793.88"), Decimal("0.09")), 
    (Decimal("4190.83"), Decimal("0.12")), 
    (Decimal("8157.41"), Decimal("0.14"))
]



def render_public_form():
    query_params = st.query_params
    if "tela" in query_params and query_params["tela"] == "admissao":
        st.title("🏗️ Admissão Inteligente - Vicelos")
        st.info("Preencha os dados abaixo. Os seus documentos são usados apenas para gerar e assinar o kit de admissão." \
                "Os campos marcados com (*) São de preenchimento obrigatório. Após clicar em Validar e Gerar Kit Admissional" \
                " você receberá um link no Whatspp para realizar a assinatura Digital dos documentos")

        REQUIRED_UPLOADS = {
            "Documento de Identidade (RG ou CNH)": "rg_cnh",
            "CPF": "cpf",
            "Comprovante de Endereço": "endereco",
            "Certidão de Nascimento ou Casamento": "certidao"
        }

        OPTIONAL_UPLOADS = {
            "CTPS (frente/verso ou digital)": "ctps",
            "Título de Eleitor": "titulo",
            "Certificado de Reservista (se houver)": "reservista",
        }

        with st.form("form_admissao", clear_on_submit=False):
            st.subheader("Dados Pessoais")
            c1, c2, c3, c4 = st.columns(4)
            nome = c1.text_input("Nome Completo *")
            dt_nasc = c2.date_input("Data de Nascimento", min_value=date(1950, 1, 1), max_value=date.today(), format="DD/MM/YYYY", value=None)
            nacionalidade = c3.text_input("Nacionalidade", value="Brasileira")
            local_nasc = c4.text_input("Local de Nascimento (Cidade/UF)")

            c5, c6, c7 = st.columns(3)
            estado_civil = c5.selectbox("Estado Civil", ["Solteiro(a)", "Casado(a)", "Divorciado(a)", "Viúvo(a)"], index=None)
            raca = c6.selectbox("Raça/Cor", ["Branca", "Parda", "Preta", "Amarela", "Indígena", "Prefiro não informar"], index=None)
            celular = c7.text_input("Celular *")

            c8, c9, c10 = st.columns(3)
            telefone = c8.text_input("Telefone Fixo")
            email = c9.text_input("E-mail *")
            pis = c10.text_input("PIS *", help="O número do PIS pode ser encontrado na sua CTPS digital.")

            st.markdown("---")
            st.subheader("Documentos")
            d1, d2, d3, d4 = st.columns(4)
            rg = d1.text_input("RG *")
            dt_rg = d2.date_input("Data de Expedição RG", min_value=date(1950, 1, 1), max_value=date.today(), format="DD/MM/YYYY", value=None)
            orgao_rg = d3.text_input("Órgão Emissor - UF (RG)")
            uf_rg = d4.text_input("UF (RG)")

            d5, d6, d7 = st.columns(3)
            cpf = d5.text_input("CPF *")
            titulo = d6.text_input("Título de Eleitor ")
            zona = d7.text_input("Zona")

            d8, d9, d10 = st.columns(3)
            secao = d8.text_input("Seção")
            mun_titulo = d9.text_input("Município/UF (Título)")
            reservista = d10.text_input("Reservista ")

            st.caption("Nota: Os dados de CTPS são extraídos automaticamente via CPF (CTPS Digital).")

            st.markdown("---")
            st.subheader("Endereço")
            e1, e2, e3, e4 = st.columns([2, 1, 1, 1])
            logradouro = e1.text_input("Endereço (Rua/Av) *")
            numero = e2.text_input("Número")
            complemento = e3.text_input("Complemento")
            bairro = e4.text_input("Bairro")
            e5, e6, e7 = st.columns([2, 1, 1])
            cidade = e5.text_input("Cidade")
            uf = e6.text_input("UF")
            cep = e7.text_input("CEP")

            st.markdown("---")
            st.subheader("Família e Dependentes")
            st.caption("Preencha os dados do cônjuge apenas se tiver certidão de casamento ou averbação. Os dependentes apenas menores de 14 anos ou com deficiência comprovada.")
            f1, f2, f3 = st.columns(3)
            conjuge = f1.text_input("Nome do Cônjuge")
            conjuge_cpf = f2.text_input("CPF do Cônjuge")
            conjuge_dt = f3.date_input("Data de Nasc. Cônjuge", min_value=date(1950, 1, 1), max_value=date.today(), format="DD/MM/YYYY", value=None)

            dep1, dep2, dep3 = st.columns(3)
            dep1_nome = dep1.text_input("Dependente 1 Nome")
            dep1_dt = dep2.date_input("Dependente 1 Data Nasc", min_value=date(1950, 1, 1), max_value=date.today(), format="DD/MM/YYYY", value=None)
            dep1_parent = dep3.text_input("Dependente 1 Parentesco")

            st.markdown("---")
            st.subheader("Dados Bancários")
            st.caption("Preencha apenas quando a sua conta no Banco Inter estiver aberta ou caso já possua conta.")
            b1, b2, b3 = st.columns(3)
            banco = b1.text_input("Banco *", value="077 - Banco Inter", disabled=True)
            agencia = b2.text_input("Agência *", value="0001", disabled=True)
            conta = b3.text_input("Conta *")
            b3.caption("Apenas dígitos (sem pontos ou traços)")

            st.markdown("---")
            st.subheader("Transporte")
            st.caption("⚖️ *Base Legal: Conforme Lei nº 7.418/1985 e Decreto nº 95.247/1987, a opção pelo Vale Transporte autoriza o desconto em folha de até 6% do seu salário base.*")
            
            vt_optin = st.radio("Deseja receber Vale Transporte? *", ["Sim", "Não"], horizontal=True, index=None)
            
            i1, i2 = st.columns(2)
            linha_ida = i1.text_input("Linha de Transporte (Ida)")
            valor_ida = i2.text_input("Valor da Passagem (Ida)")
            v1, v2 = st.columns(2)
            linha_volta = v1.text_input("Linha de Transporte (Volta)")
            valor_volta = v2.text_input("Valor da Passagem (Volta)")

            st.markdown("---")
            st.subheader("Uploads Obrigatórios")
            uploads = {}
            for label, key in REQUIRED_UPLOADS.items():
                file = st.file_uploader(label, type=["pdf", "png", "jpg", "jpeg"], key=key)
                uploads[key] = file

            st.subheader("Uploads Opcionais")
            for label, key in OPTIONAL_UPLOADS.items():
                file = st.file_uploader(label, type=["pdf", "png", "jpg", "jpeg", "zip"], key=key)
                uploads[key] = file

            st.markdown("---")
            st.subheader("Finalização")
            
            declaracao = st.checkbox(
                "Declaro, para os devidos fins de direito, que todas as informações aqui prestadas são verdadeiras," \
                " completas e autênticas, assumindo inteira responsabilidade legal pela veracidade das mesmas." \
                " Estou ciente de que a omissão ou a falsidade de dados constitui infração legal," \
                " sujeitando-me às sanções cabíveis e à rescisão do processo de admissão." \
                " Autorizo, nos termos da LGPD (Lei nº 13.709/2018), o tratamento dos meus dados e documentos " \
                "para fins exclusivos de recrutamento, seleção e registro de contrato de trabalho pela Vicelos."
            )

            submitted = st.form_submit_button("Validar e GERAR KIT DE ADMISSÃO")

        if submitted:
            missing_files = [lbl for lbl, k in REQUIRED_UPLOADS.items() if uploads.get(k) is None]
            if not nome or not cpf or not email or not pis or not celular or not rg or not logradouro or not bairro or not cidade  or not numero or not cidade or not cep or vt_optin is None:
                st.error("🚨 Por favor, preencha os campos obrigatórios básicos (Nome, CPF, E-mail, Telefone , Endereço , CEP , PIS e opção de VT).")
            elif missing_files:
                st.error(f"🚨 Por favor, anexe os documentos obrigatórios: {', '.join(missing_files)}")
            elif not declaracao:
                st.error("🚨 Você precisa marcar a declaração de veracidade antes de enviar.")
            else:
                st.success(f"Obrigado, {nome}! Gerando documentos e planilha...")

                dados_finais = {
                    'nome': nome,
                    'cpf': cpf,
                    'rg': rg,
                    'pis': pis,
                    'cargo': '',
                    'cbo': '',
                    'salario': '0,00',
                    'VT Diário': valor_ida or "0,00",
                    'data_inicio': datetime.now().strftime('%d/%m/%Y'),
                    'obra': '',
                    'nacionalidade': nacionalidade,
                    'estado_civil': estado_civil,
                    'Logradouro': logradouro,
                    'Numero Endereco': numero,
                    'Bairro': bairro,
                    'Cidade': cidade,
                    'Estado': uf,
                    'CEP': cep,
                    'Complemento': complemento,
                    'Data de Nascimento': dt_nasc.strftime('%d/%m/%Y') if dt_nasc else '',
                    'Local de Nascimento': local_nasc,
                    'Orgao Emissor RG': orgao_rg,
                    'UF RG': uf_rg,
                    'Titulo de Eleitor': titulo,
                    'Zona': zona,
                    'Secao': secao,
                    'Reservista': reservista,
                    'mae': '', 'pai': '',
                    'email': email,
                    'telefone': telefone,
                    'celular': celular,
                    'vt_optin': vt_optin,
                    'linha_ida': linha_ida,
                    'valor_ida': valor_ida,
                    'linha_volta': linha_volta,
                    'valor_volta': valor_volta,
                }

                with st.spinner("Gerando kit no Drive e preenchendo planilha..."):
                    link_pasta, pasta_id = gerar_kit_admissional(dados_finais)

                with st.spinner("Finalizando processo..."):
                    st.success("Obrigado, seu processo foi concluído!")
                    st.markdown(f"**Pasta no Drive:** [abrir]({link_pasta})")

        return

# ==========================================
# 4. FUNÇÕES DE INFRAESTRUTURA E RESTO DO CÓDIGO
# ==========================================
# (Seu código já começa a partir daqui com a função conectar_google)

def conectar_google():
    SCOPES = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive', 'https://www.googleapis.com/auth/documents']
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    token_pickle = os.path.join(BASE_DIR, 'token.json')
    client_secret_json = os.path.join(BASE_DIR, 'client_secret.json')

    creds = None
    if os.path.exists(token_pickle):
        creds = Credentials.from_authorized_user_file(token_pickle, SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            try:
                creds.refresh(Request())
            except google_exceptions.RefreshError:
                # Se o token expirou ou foi revogado (ex: chave deletada), removemos o arquivo para reautenticar
                if os.path.exists(token_pickle):
                    os.remove(token_pickle)
                creds = None
        
        # Se após o refresh as credenciais ainda forem inválidas ou nulas, gera um novo fluxo
        if not creds or not creds.valid:
            flow = InstalledAppFlow.from_client_secrets_file(client_secret_json, SCOPES)
            creds = flow.run_local_server(port=0)
            
        with open(token_pickle, 'w') as token:
            token.write(creds.to_json())
    return build('drive', 'v3', credentials=creds), build('docs', 'v1', credentials=creds), gspread.authorize(creds)

def formatar_moeda(valor):
    if not valor or valor == 0 or valor == "0,00": return "0,00"
    if isinstance(valor, str):
        valor = _to_decimal(valor)
    return f"{valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

# ==========================================
# 3. MÓDULO ADMISSÃO
# ==========================================

def processar_documentos_ia(arquivos_upload):
    genai.configure(api_key=API_KEY_GEMINI)
    modelo = genai.GenerativeModel('gemini-2.5-flash', generation_config={"response_mime_type": "application/json"})
    arquivos_ia = []
    status_bar = st.progress(0, text="Enviando arquivos para a IA...")
    for i, arq in enumerate(arquivos_upload):
        with open(f"temp_{arq.name}", "wb") as f: f.write(arq.getbuffer())   
        arquivo_up = genai.upload_file(path=f"temp_{arq.name}")
        while arquivo_up.state.name == "PROCESSING": time.sleep(10); arquivo_up = genai.get_file(arquivo_up.name)
        arquivos_ia.append(arquivo_up); os.remove(f"temp_{arq.name}")
        status_bar.progress((i + 1) / len(arquivos_upload), text=f"Lendo: {arq.name}")

    prompt = """Você é um sistema automatizado de RH. Sua única função é extrair dados dos documentos fornecidos e preencher o template JSON abaixo.
Os documentos podem incluir CNH, RG, CRNM de estrangeiros CTPS, Comprovante de Endereço, Título de Eleitor, Certificado de Reservista e Certidões (Nascimento/Casamento).
REGRAS:
1. "Nome Completo", "CPF", "RG", "Data de Nascimento": CNH, RG ou CRNM.
2. "Nome da Mae" e "Nome do Pai": CNH, RG, CRNM ou Certidão.
3. Endereço: Comprovante de residência. Ignore o endereço da empresa emissora (ex: Sabesp) e extraia O ENDEREÇO DO CONSUMIDOR. Tente desmembrar em Logradouro, Número, Bairro, Cidade e CEP. Se houver mais de um comprovante, escolha o mais recente.
4. "CTPS Numero": Se CTPS Digital, use o CPF. Deixe Série e UF vazios.
5. "Orgao Emissor RG" e "UF RG": Separe as letras do Estado (ex: SSP e SP). Para estrangeiros, deixe vazio.
6. "Titulo de Eleitor", "Zona", "Secao": Título de Eleitor.
7. "Reservista": Certificado militar.
8. "Estado Civil": Siga a regra documental rigorosamente:
  - Se o documento apresentado for uma "Certidão de Nascimento" (e não houver averbação de casamento), preencha "SOLTEIRO(A)".
  - Se o documento for uma "Certidão de Casamento", preencha "CASADO(A)"

    JSON ESPERADO:
    {
      "Nome Completo": "", "Data de Nascimento": "", "Local de Nascimento": "", "Estado Civil": "",
      "CPF": "", "RG": "", "Orgao Emissor RG": "", "UF RG": "", "Nome da Mae": "", "Nome do Pai": "",
      "CEP": "", "Logradouro": "", "Numero Endereco": "", "Complemento": "", "Bairro": "",
      "Cidade": "", "Estado": "", "PIS": "", "CTPS Numero": "", "CTPS Serie": "", "CTPS UF": "",
      "Titulo de Eleitor": "", "Zona": "", "Secao": "", "Reservista": ""
    }
    """
    res = modelo.generate_content([prompt] + arquivos_ia, request_options={"timeout": 600})
    return json.loads(res.text)

def gerar_kit_admissional(dados_finais):
    drive_service, docs_service, gc = conectar_google()
    salario_raw = dados_finais.get('salario', '0,00')
    try: sal_extenso = num2words(float(salario_raw.replace('.','').replace(',','.')), lang='pt_BR', to='currency')
    except: sal_extenso = "Valor não identificado"
    
    try:
        dt_inicio = datetime.strptime(dados_finais.get('data_inicio', datetime.now().strftime('%d/%m/%Y')), '%d/%m/%Y')
    except Exception:
        dt_inicio = datetime.now()
    dt_fim_45 = dt_inicio + timedelta(days=44)
    dt_fim_90 = dt_inicio + timedelta(days=89)
    
    tags = {
        '{{NOME}}': dados_finais.get('nome', '').upper(),
        '{{NACIONALIDADE}}': dados_finais.get('nacionalidade', 'Brasileiro').upper(),
        '{{ESTADO_CIVIL}}': dados_finais.get('estado_civil', dados_finais.get('Estado Civil', 'Solteiro(a)')).upper(),
        '{{PROFISSAO}}': dados_finais.get('cargo', ''),
        '{{RG}}': dados_finais.get('rg', ''),
        '{{CPF}}': dados_finais.get('cpf', ''),
        '{{PIS}}': dados_finais.get('pis', ''),
        '{{ENDERECO_COMPLETO}}': f"{dados_finais.get('Logradouro', dados_finais.get('logradouro', ''))} {dados_finais.get('Numero Endereco', dados_finais.get('numero', ''))}, {dados_finais.get('Bairro', dados_finais.get('bairro', ''))}, {dados_finais.get('Cidade', dados_finais.get('cidade', ''))}-{dados_finais.get('Estado', dados_finais.get('estado', ''))}",
        '{{CTPS_NUM}}': dados_finais.get('CTPS Numero', dados_finais.get('ctps_num', '')),
        '{{CTPS_SERIE}}': dados_finais.get('CTPS Serie', dados_finais.get('ctps_serie', '')),
        '{{FUNCAO}}': dados_finais.get('cargo', ''),
        '{{SALARIO}}': dados_finais.get('salario', ''),
        '{{SALARIO_EXTENSO}}': sal_extenso,
        '{{DATA_INICIO}}': dados_finais.get('data_inicio', ''),
        '{{FIM_45}}': dt_fim_45.strftime('%d/%m/%Y'),
        '{{FIM_90}}': dt_fim_90.strftime('%d/%m/%Y'),
        '{{DATA_HOJE}}': f"São Paulo, {datetime.now().day} de {datetime.now().strftime('%B')} de {datetime.now().year}"
    }

    metadata = {'name': f"Admissão - {tags['{{NOME}}']}", 'mimeType': 'application/vnd.google-apps.folder', 'parents': [ID_PASTA_RAIZ]}
    pasta = drive_service.files().create(body=metadata, fields='id').execute()
    id_pasta = pasta.get('id'); link_pasta = f"https://drive.google.com/drive/folders/{id_pasta}"

    progresso = st.progress(0); contador = 0
    for nome_doc, id_modelo in MODELOS_ADMISSAO.items():
        copia = drive_service.files().copy(fileId=id_modelo, body={'name': f"{nome_doc} - {tags['{{NOME}}']}", 'parents': [id_pasta]}).execute()
        requests = [{'replaceAllText': {'containsText': {'text': k, 'matchCase': True}, 'replaceText': str(v)}} for k, v in tags.items()]
        docs_service.documents().batchUpdate(documentId=copia.get('id'), body={'requests': requests}).execute()
        contador += 1; progresso.progress(contador / len(MODELOS_ADMISSAO))

    aba = gc.open_by_key(ID_PLANILHA).worksheet('Base_de_Dados_Funcionarios')
    col_mat = aba.col_values(1)
    maior = 0
    for v in col_mat[1:]:
        if v.isdigit() and int(v) > maior: maior = int(v)
    nova_mat = f"{(maior + 1):06d}"

    linha = [""] * 65
    linha[0] = nova_mat; linha[1] = nova_mat; linha[2] = tags['{{NOME}}']
    linha[3] = dados_finais.get('Data de Nascimento', dados_finais.get('nascimento', ''))
    linha[4] = tags['{{NACIONALIDADE}}']
    linha[5] = dados_finais.get('Local de Nascimento', dados_finais.get('local_nasc', ''))
    linha[6] = tags['{{ESTADO_CIVIL}}']
    linha[8] = tags['{{RG}}']
    linha[10] = dados_finais.get('Orgao Emissor RG', dados_finais.get('orgao_rg', ''))
    linha[11] = dados_finais.get('UF RG', dados_finais.get('uf_rg', ''))
    linha[12] = tags['{{CPF}}']; linha[13] = tags['{{PIS}}']; 
    linha[14] = tags['{{CTPS_NUM}}']; linha[15] = tags['{{CTPS_SERIE}}']
    linha[16] = dados_finais.get('CTPS UF', dados_finais.get('ctps_uf', ''))
    linha[18] = dados_finais.get('Titulo de Eleitor', dados_finais.get('titulo', ''))
    linha[19] = dados_finais.get('Zona', dados_finais.get('zona', ''))
    linha[20] = dados_finais.get('Secao', dados_finais.get('secao', ''))
    linha[22] = dados_finais.get('Reservista', dados_finais.get('reservista', ''))
    linha[23] = dados_finais.get('Nome da Mae', dados_finais.get('mae', ''))
    linha[24] = dados_finais.get('Nome do Pai', dados_finais.get('pai', ''))
    linha[25] = dados_finais.get('Logradouro', dados_finais.get('logradouro', ''))
    linha[26] = dados_finais.get('Numero Endereco', dados_finais.get('numero', ''))
    linha[27] = dados_finais.get('Complemento', dados_finais.get('complemento', ''))
    linha[28] = dados_finais.get('Bairro', dados_finais.get('bairro', ''))
    linha[29] = dados_finais.get('Cidade', dados_finais.get('cidade', ''))
    linha[30] = dados_finais.get('Estado', dados_finais.get('estado', ''))
    linha[31] = dados_finais.get('CEP', dados_finais.get('cep', ''))
    linha[53] = dados_finais.get('VT Diário', '0,00') 
    linha[54] = tags['{{DATA_INICIO}}']; linha[55] = tags['{{FUNCAO}}']
    linha[56] = dados_finais.get('cbo', '')
    linha[57] = tags['{{SALARIO}}']; linha[60] = link_pasta
    linha[64] = dados_finais.get('obra', '') 
    
    aba.append_row(linha)
    return link_pasta, id_pasta


def exportar_pdfs_da_pasta(folder_id: str):
    drive_service, _, _ = conectar_google()
    res = drive_service.files().list(q=f"'{folder_id}' in parents and mimeType='application/vnd.google-apps.document'",
                                     fields="files(id, name)").execute()
    pdfs = []
    for f in res.get('files', []):
        request = drive_service.files().export_media(fileId=f['id'], mimeType='application/pdf')
        pdf_bytes = request.execute()
        pdfs.append((f"{f['name']}.pdf", pdf_bytes))
    return pdfs

# ==========================================================
# 3. MÓDULO PÚBLICO: PORTAL DO CANDIDATO (FORMULÁRIO COMPLETO)
# ==========================================================

# ==========================================
# 4. MÓDULO FOLHA
# ==========================================

def calcular_inss(bruto):
    if not isinstance(bruto, Decimal): bruto = _to_decimal(bruto)
    desconto = Decimal("0.00")
    faixa_ant = Decimal("0.00")
    for teto, aliq in FAIXAS_INSS:
        if bruto > faixa_ant:
            base = min(bruto, teto) - faixa_ant
            desconto += base * aliq
            faixa_ant = teto
        else: break
    return desconto.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

def salvar_log_folha_sheets(lista_dados):
    """Salva um resumo do processamento de folha em uma aba do Google Sheets para conferência."""
    if not lista_dados: return
    try:
        # Importar internamente para evitar dependências circulares ou pesos desnecessários
        import gspread
        _, _, gc = conectar_google()
        planilha = gc.open_by_key(ID_PLANILHA)
        
        try:
            aba = planilha.worksheet('Log_Processamento_Folha')
        except:
            aba = planilha.add_worksheet(title='Log_Processamento_Folha', rows="500", cols="30")
            aba.append_row([
                "Data Proc.", "Competência", "CPF", "Nome", 
                "Sal. Base", "HE+DSR", "Bruto", 
                "INSS", "IRRF", "VT", "VA", "Sind.", "Faltas", "Atrasos", "Adiant.", 
                "Tot. Desc.", "Líquido", "FGTS", "Banco"
            ])

        rows_to_append = []
        for d in lista_dados:
            rows_to_append.append([
                datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
                d.get('competencia', ''),
                d.get('cpf', ''),
                d.get('nome', ''),
                round(float(d.get('salario_base', 0)), 2),
                round(float(d.get('he_total', 0)), 2),
                round(float(d.get('bruto', 0)), 2),
                round(float(d.get('inss', 0)), 2),
                round(float(d.get('irrf', 0)), 2),
                round(float(d.get('vt', 0)), 2),
                round(float(d.get('va', 0)), 2),
                round(float(d.get('sindical', 0)), 2),
                round(float(d.get('faltas', 0)), 2),
                round(float(d.get('atrasos', 0)), 2),
                round(float(d.get('adiantamento', 0)), 2),
                round(float(d.get('desconto_total', 0)), 2),
                round(float(d.get('liquido', 0)), 2),
                round(float(d.get('fgts', 0)), 2),
                d.get('banco', '')
            ])
        
        aba.append_rows(rows_to_append)
        return True
    except Exception as e:
        st.error(f"Erro ao salvar log no Sheets: {e}")
        return False

def calcular_irrf_2026(bruto, desconto_inss, dependentes=0):
    if not isinstance(bruto, Decimal): bruto = _to_decimal(bruto)
    if not isinstance(desconto_inss, Decimal): desconto_inss = _to_decimal(desconto_inss)
    
    deducao_dep = Decimal("189.59") * Decimal(str(dependentes))
    base_legal = bruto - desconto_inss - deducao_dep
    base_simplificada = bruto - Decimal("607.20")
    base_calculo = min(base_legal, base_simplificada)
    
    # Tabela Tradicional 2026 (Baseada no PDF)
    if base_calculo <= Decimal("2428.80"): 
        imposto_bruto = Decimal("0.00")
        faixa_str = "ISENTO"
    elif base_calculo <= Decimal("2826.65"): 
        imposto_bruto = (base_calculo * Decimal("0.075")) - Decimal("182.16")
        faixa_str = "7.5%"
    elif base_calculo <= Decimal("3751.05"): 
        imposto_bruto = (base_calculo * Decimal("0.150")) - Decimal("394.16")
        faixa_str = "15.0%"
    elif base_calculo <= Decimal("4664.68"): 
        imposto_bruto = (base_calculo * Decimal("0.225")) - Decimal("675.49")
        faixa_str = "22.5%"
    else: 
        imposto_bruto = (base_calculo * Decimal("0.275")) - Decimal("908.73")
        faixa_str = "27.5%"
    
    # Regra de Redução Adicional (Nova Tabela 2026)
    # Redução para quem ganha até R$ 5.000 (Zera o imposto)
    if bruto <= Decimal("5000.00"): 
        return Decimal("0.00"), "ISENTO (Redutor 2026)"
    # Redução Gradual para quem ganha até R$ 7.350
    elif bruto <= Decimal("7350.00"):
        reducao = Decimal("978.62") - (Decimal("0.133145") * bruto)
        imposto_final = max(Decimal("0.00"), imposto_bruto - reducao)
        return imposto_final.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP), (faixa_str if imposto_final > 0 else "ISENTO")
    
    return imposto_bruto.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP), faixa_str

def calcular_adiantamento_prop(sal_base, adm_str, fecham_str):
    if not isinstance(sal_base, Decimal): sal_base = _to_decimal(sal_base)
    try:
        dt_adm = datetime.strptime(adm_str, '%d/%m/%Y')
        dt_fech = datetime.strptime(fecham_str, '%d/%m/%Y')

        # Passo A: Valor do Dia Comercial
        valor_dia = sal_base / Decimal("30")

        # Passo B: Dias de Direito no mês de admissão (mês comercial de 30 dias)
        if dt_adm.year == dt_fech.year and dt_adm.month == dt_fech.month:
            dias_trabalhados = max(0, 30 - dt_adm.day + 1)
            salario_prop = valor_dia * Decimal(str(dias_trabalhados))
            return (salario_prop * Decimal("0.40")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

        return (sal_base * Decimal("0.40")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except Exception:
        return (sal_base * Decimal("0.40")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

def gerar_holerite_dinamico(dados):
    drive, docs, _ = conectar_google()
    salario_integral = _to_decimal(dados['salario_base'])
    tipo_processamento = dados.get('tipo_processamento', 'Fechamento Mensal')

    if tipo_processamento == "Adiantamento Quinzenal (Dia 20)":
        bruto = calcular_adiantamento_prop(salario_integral, dados['admissao'], dados['data_fechamento'])
        liquido = bruto; total_desc = Decimal("0.00"); fgts = Decimal("0.00"); inss = Decimal("0.00"); faixa_irrf = ""

        try:
            dt_adm = datetime.strptime(dados['admissao'], '%d/%m/%Y')
            dt_fech = datetime.strptime(dados['data_fechamento'], '%d/%m/%Y')
            if dt_adm.year == dt_fech.year and dt_adm.month == dt_fech.month:
                dias_trab = max(0, 30 - dt_adm.day + 1)
                base_prop = (salario_integral / Decimal("30")) * Decimal(str(dias_trab))
            else:
                base_prop = salario_integral
        except Exception:
            base_prop = salario_integral

        fator_texto = f"{(bruto / base_prop * 100):.2f}" if base_prop > 0 else "0.00"
        rubricas = [["001", "ADIANTAMENTO QUINZENAL", fator_texto, formatar_moeda(bruto), ""]]
    else:
        try:
            dt_admissao = datetime.strptime(dados['admissao'], '%d/%m/%Y')
            dt_fechamento = datetime.strptime(dados['data_fechamento'], '%d/%m/%Y')
            # Mês comercial 30 dias se admitido em meses anteriores
            if dt_admissao.month == dt_fechamento.month and dt_admissao.year == dt_fechamento.year:
                dias_trabalhados = (dt_fechamento - dt_admissao).days + 1
            else:
                dias_trabalhados = 30
        except: dias_trabalhados = 30
            
        salario_proporcional = (salario_integral / Decimal("30")) * Decimal(str(dias_trabalhados))
        val_dia = salario_integral / Decimal("30")
        val_hora = salario_integral / Decimal("220")
        
        he_val = (val_hora * Decimal("1.6")) * Decimal(str(dados.get('qtd_he', 0.0)))
        he_100_val = (val_hora * Decimal("2.0")) * Decimal(str(dados.get('qtd_he_100', 0.0)))
        dsr_val = ((he_val + he_100_val) / Decimal("25")) * Decimal("5") if (he_val + he_100_val) > 0 else Decimal("0.00")
        
        # 1. Soma dos Vencimentos (Bruto)
        bruto = salario_proporcional + he_val + he_100_val + dsr_val
        
        # 2. Desconto de Faltas, DSR e Atrasos
        val_faltas = Decimal(str(dados.get('dias_faltas', 0))) * val_dia
        val_dsr_perdido = Decimal(str(dados.get('dsr_descontado', 0))) * val_dia
        val_atrasos = Decimal(str(dados.get('horas_atrasos', 0.0))) * val_hora
        
        # 3. Base Real de Impostos (O que sobra depois das faltas)
        base_impostos = bruto - val_faltas - val_dsr_perdido - val_atrasos
        if base_impostos < 0: base_impostos = Decimal("0.00")
        
        # 4. Cálculo dos Impostos sobre a Base Real
        inss = calcular_inss(base_impostos)
        irrf, faixa_irrf = calcular_irrf_2026(base_impostos, inss, dependentes=dados.get('dependentes', 0))
        
        # 5. Novo Motor de Vale Transporte
        desc_vt_flag = dados.get('Descontar VT', False)
        dias_uteis_vt = Decimal(str(dados.get('Dias Úteis VT', 22)))
        custo_diario_vt = _to_decimal(dados.get('VT Diário', '0'))
            
        custo_total_vt = custo_diario_vt * dias_uteis_vt
        limite_6_pct = salario_proporcional * Decimal("0.06")
        vt = min(custo_total_vt, limite_6_pct) if (custo_total_vt > 0 and desc_vt_flag) else Decimal("0.00")
        
        # 6. Outros Descontos
        adiantamento = calcular_adiantamento_prop(salario_integral, dados['admissao'], dados['data_fechamento']) if dados.get('Desc. Adiant.?', False) else Decimal("0.00")
        
        if dados.get('Descontar VA', False):
            # Usar o valor do banco se existir, ou o proporcional de 485,00
            va_cheio = _to_decimal(dados.get('Valor VA', '0'))
            if va_cheio <= 0: va_cheio = Decimal("485.00")
            # O desconto é de 5% sobre o valor do benefício (que por sua vez é proporcional aos dias)
            beneficio_prop = ((va_cheio / Decimal("30")) * Decimal(str(dias_trabalhados))).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            desc_cesta = (beneficio_prop * Decimal("0.05")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        else:
            desc_cesta = Decimal("0.00")
            
        desc_sindical = min(base_impostos * Decimal("0.01"), Decimal("46.30")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP) if str(dados.get('oposicao', '')).strip().upper() != 'SIM' else Decimal("0.00")
        
        # 7. Totalizadores
        total_desc = inss + vt + irrf + adiantamento + val_faltas + val_dsr_perdido + val_atrasos + desc_cesta + desc_sindical
        liquido = bruto - total_desc
        fgts = (base_impostos * Decimal("0.08")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        
        rubricas = [["001", "SALARIO BASE", f"{dias_trabalhados} .00", formatar_moeda(salario_proporcional), ""]]
        if he_val > 0: rubricas.append(["002", "HORAS EXTRAS 60%", f"{dados.get('qtd_he')}", formatar_moeda(he_val), ""])
        if he_100_val > 0: rubricas.append(["004", "HORAS EXTRAS 100%", f"{dados.get('qtd_he_100')}", formatar_moeda(he_100_val), ""])
        if dsr_val > 0: rubricas.append(["003", "D.S.R. S/ VARIAVEIS", "", formatar_moeda(dsr_val), ""])
        if inss > 0: rubricas.append(["101", "INSS", f"{(inss/base_impostos*100):.2f}" if base_impostos > 0 else "0.00", "", formatar_moeda(inss)])
        if irrf > 0: rubricas.append(["102", "IRRF", faixa_irrf, "", formatar_moeda(irrf)])
        if vt > 0: rubricas.append(["103", "VALE TRANSPORTE", "6,00 %", "", formatar_moeda(vt)])
        if adiantamento > 0: rubricas.append(["104", "ADIANTAMENTO", "", "", formatar_moeda(adiantamento)])
        if desc_cesta > 0: rubricas.append(["107", "VALE ALIMENTAÇÃO", "5.00", "", formatar_moeda(desc_cesta)])
        if desc_sindical > 0: rubricas.append(["108", "CONTRIB. SINDICAL", "1.00", "", formatar_moeda(desc_sindical)])
        if val_faltas > 0: rubricas.append(["105", "FALTAS (DIAS)", f"{dados['dias_faltas']}", "", formatar_moeda(val_faltas)])
        if val_dsr_perdido > 0: rubricas.append(["109", "PERDA DSR (FALTAS)", f"{dados.get('dsr_descontado')}", "", formatar_moeda(val_dsr_perdido)])
        if val_atrasos > 0: rubricas.append(["106", "ATRASOS (HORAS)", f"{dados['horas_atrasos']}", "", formatar_moeda(val_atrasos)])

    tags = {
        '{{NOME}}': dados['nome'], '{{CARGO}}': dados['cargo'], '{{CBO}}': dados.get('cbo', ''),
        '{{MES_REFERENCIA}}': datetime.strptime(dados['data_fechamento'], '%d/%m/%Y').strftime('%m/%Y'), '{{DATA_HOJE}}': datetime.now().strftime('%d/%m/%Y'),
        '{{OBRA}}': dados['obra'], '{{DATA_INICIO}}': dados.get('admissao', ''), '{{SALARIO}}': formatar_moeda(salario_integral), 
        '{{TOTAL_VENC}}': formatar_moeda(bruto), '{{TOTAL_DESC}}': formatar_moeda(total_desc), '{{VALOR_LIQUIDO}}': formatar_moeda(liquido),
        '{{VALOR_EXTENSO}}': num2words(float(liquido), lang='pt_BR', to='currency'), 
        
        '{{BASE_INSS}}': formatar_moeda(base_impostos if tipo_processamento != "Adiantamento Quinzenal (Dia 20)" else Decimal("0.00")), 
        '{{BASE_FGTS}}': formatar_moeda(base_impostos if tipo_processamento != "Adiantamento Quinzenal (Dia 20)" else Decimal("0.00")), 
        '{{VALOR_FGTS}}': formatar_moeda(fgts), 
        '{{BASE_IRRF}}': formatar_moeda((base_impostos - inss) if tipo_processamento != "Adiantamento Quinzenal (Dia 20)" else Decimal("0.00")), 
        '{{FAIXA_IRRF}}': faixa_irrf
    }
    for i in range(12):
        num = f"{i+1:02d}"; r = rubricas[i] if i < len(rubricas) else ["","","","",""]
        tags[f'{{{{C{num}}}}}'] = r[0]; tags[f'{{{{D{num}}}}}'] = r[1]; tags[f'{{{{R{num}}}}}'] = r[2]; tags[f'{{{{V{num}}}}}'] = r[3]; tags[f'{{{{DE{num}}}}}'] = r[4]

    copia = drive.files().copy(fileId=ID_MODELO_HOLERITE, body={'name': f"Holerite - {dados['nome']}", 'parents': [ID_PASTA_RAIZ]}).execute()
    reqs = [{'replaceAllText': {'containsText': {'text': k, 'matchCase': True}, 'replaceText': str(v)}} for k, v in tags.items()]
    docs.documents().batchUpdate(documentId=copia.get('id'), body={'requests': reqs}).execute()
    
    detalhes = {
        'salario_base': salario_proporcional if tipo_processamento != "Adiantamento Quinzenal (Dia 20)" else bruto,
        'he_60': he_val if tipo_processamento != "Adiantamento Quinzenal (Dia 20)" else Decimal("0.00"),
        'he_100': he_100_val if tipo_processamento != "Adiantamento Quinzenal (Dia 20)" else Decimal("0.00"),
        'dsr': dsr_val if tipo_processamento != "Adiantamento Quinzenal (Dia 20)" else Decimal("0.00"),
        'bruto': bruto,
        'inss': inss,
        'irrf': irrf,
        'vt': vt,
        'va': desc_cesta,
        'adiantamento': adiantamento if tipo_processamento != "Adiantamento Quinzenal (Dia 20)" else Decimal("0.00"),
        'desc_total': total_desc,
        'liquido': liquido,
        'faltas': val_faltas + val_dsr_perdido,
        'atrasos': val_atrasos,
        'sindical': desc_sindical,
        'base_calc': base_impostos,
        'fgts': fgts
    }
    return f"https://docs.google.com/document/d/{copia.get('id')}/edit", liquido, detalhes

def gerar_excel_banco_inter(lista_pagamentos, total_folha):
    import openpyxl
    from openpyxl.styles import NamedStyle
    
    template_path = "/Users/renancarvalho/Downloads/Template_Folha_de_Pagamento.xlsx"
    total_real = round(sum(p['Valor'] for p in lista_pagamentos), 2)
    
    # Tentamos carregar o template oficial para garantir 100% de compatibilidade
    if os.path.exists(template_path):
        try:
            wb = openpyxl.load_workbook(template_path)
            
            # 1. Aba Pagador
            ws_p = wb['Pagador']
            # De acordo com o template e o erro, os dados devem estar na linha 2
            ws_p['A2'] = 'VICELOS ENGENHARIA E CONSTRUÇÃO LTDA.'
            ws_p['B2'] = '37742513000188'
            ws_p['C2'] = '69655910'
            ws_p['D2'] = total_real
            # Forçamos o formato numérico do template se necessário, mas o WB já deve ter
            ws_p['D2'].number_format = '"R$"\ #,##0.00'
            
            # 2. Aba Beneficiarios
            ws_b = wb['Beneficiarios']
            # Limpamos dados antigos se houver (o template costuma vir limpo da linha 2 em diante)
            for row in range(2, ws_b.max_row + 1):
                for col in range(1, 7):
                    ws_b.cell(row=row, column=col).value = None
            
            for i, p in enumerate(lista_pagamentos):
                row = i + 2
                cpf_limpo = str(p['CPF']).replace('.', '').replace('-', '').strip()
                # Preservamos o hífen na conta, pois o banco pode exigir o dígito separado
                conta_original = str(p['Conta']).strip()
                conta_limpa = conta_original.replace('.', '')
                
                dt_debito = datetime.strptime(p['Data Debito'], "%d/%m/%Y")
                dt_pagto = datetime.strptime(p['Data Pagamento'], "%d/%m/%Y")
                
                ws_b.cell(row=row, column=1, value=str(p['Nome']).strip())
                ws_b.cell(row=row, column=2, value=cpf_limpo)
                ws_b.cell(row=row, column=3, value=conta_limpa)
                ws_b.cell(row=row, column=4, value=float(p['Valor']))
                ws_b.cell(row=row, column=5, value=dt_debito)
                ws_b.cell(row=row, column=6, value=dt_pagto)
                
                # Aplicamos formatos idênticos ao template
                ws_b.cell(row=row, column=2).number_format = '@' # CPF Texto
                ws_b.cell(row=row, column=3).number_format = '@' # Conta Texto
                ws_b.cell(row=row, column=4).number_format = '"R$"\ #,##0.00'
                ws_b.cell(row=row, column=5).number_format = 'dd/mm/yyyy'
                ws_b.cell(row=row, column=6).number_format = 'dd/mm/yyyy'

            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()
        except Exception as e:
            st.error(f"Erro ao usar template: {e}. Usando gerador fallback...")

    # FALLBACK: Se o template não existir, usamos o xlsxwriter (tentando ser o mais fiel possível)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        pd.DataFrame([['1- Não altere...'], ['...']]).to_excel(writer, sheet_name='Instruções', index=False, header=False)
        
        df_pagador = pd.DataFrame([
            ['VICELOS ENGENHARIA E CONSTRUÇÃO LTDA.', '37742513000188', '69655910', total_real]
        ], columns=['Empresa', 'CNPJ', 'Conta Corrente', 'Valor Total da Folha de Pagamento'])
        df_pagador.to_excel(writer, sheet_name='Pagador', index=False)
        
        dados_b = []
        for p in lista_pagamentos:
            dados_b.append([
                str(p['Nome']).strip(),
                str(p['CPF']).replace('.', '').replace('-', '').strip(),
                str(p['Conta']).replace('.', '').replace('-', '').strip(),
                float(p['Valor']),
                datetime.strptime(p['Data Debito'], "%d/%m/%Y"),
                datetime.strptime(p['Data Pagamento'], "%d/%m/%Y")
            ])
        df_b = pd.DataFrame(dados_b, columns=['Nome','CPF','Conta Corrente - Receb','Valor','Data do Débito','Data do Pagamento'])
        df_b.to_excel(writer, sheet_name='Beneficiarios', index=False)
        
        workbook = writer.book
        fmt_val = workbook.add_format({'num_format': '"R$"\ #,##0.00'})
        fmt_txt = workbook.add_format({'num_format': '@'})
        fmt_dat = workbook.add_format({'num_format': 'dd/mm/yyyy'})
        
        ws_b = writer.sheets['Beneficiarios']
        ws_b.set_column('B:C', None, fmt_txt)
        ws_b.set_column('D:D', None, fmt_val)
        ws_b.set_column('E:F', None, fmt_dat)
        
        ws_p = writer.sheets['Pagador']
        ws_p.set_column('B:C', None, fmt_txt)
        ws_p.set_column('D:D', None, fmt_val)

    return output.getvalue()

def buscar_funcionarios():
    # 1. Tentar Postgres (Prioritário)
    lista_db = []
    try:
        lista_db = get_funcionarios_financeiro()
    except Exception as e:
        st.warning(f"Erro ao buscar funcionários no Banco de Dados: {e}")

    # 2. Tentar Google Sheets (Fallback/Legado)
    lista_sheets = []
    try:
        _, _, gc = conectar_google()
        aba = gc.open_by_key(ID_PLANILHA).worksheet('Base_de_Dados_Funcionarios')
        dados = aba.get_all_values()
        for linha in dados[1:]:
            if len(linha) > 50 and linha[2] and (linha[62] if len(linha) > 62 else "ATIVO").strip().upper() != "INATIVO":
                lista_sheets.append({
                    'nome': linha[2], 'cpf': linha[12], 'admissao': linha[54], 
                    'cargo': linha[55], 'cbo': linha[56], 'salario': linha[57], 
                    'banco': linha[46] if len(linha) > 46 else "",
                    'conta': linha[48], 'oposicao': linha[61], 'obra': linha[64],
                    'VT Diário': linha[52] if len(linha) > 52 else "0,00",
                    'Valor VA': linha[58] if len(linha) > 58 else "0,00",
                    'Valor VR': linha[59] if len(linha) > 59 else "0,00",
                    'Descontar VT': True,
                    'Descontar VA': True
                })
    except Exception as e:
        st.warning(f"Erro ao buscar funcionários no Google Sheets: {e}")

    # 3. Mesclagem Inteligente (Merge por CPF)
    # Se o CPF já existe no DB, preferimos os dados do DB (que tendem a estar mais atualizados com o ERP)
    cpfs_db = {f['cpf'].replace('.','').replace('-','').strip() for f in lista_db}
    
    final_list = list(lista_db)
    for fs in lista_sheets:
        cpf_s = fs['cpf'].replace('.','').replace('-','').strip()
        if cpf_s not in cpfs_db:
            final_list.append(fs)
            
    return final_list

# ==========================================
# NOVO MÓDULO: SIMULADOR DE DESLIGAMENTO
# ==========================================

def calcular_cenarios_desligamento(func_dados, data_desligamento):
    try: dt_adm = datetime.strptime(func_dados['admissao'], '%d/%m/%Y').date()
    except Exception as e: return None
    
    salario = _to_decimal(func_dados.get('salario', '0'))
    va_mensal = _to_decimal(func_dados.get('Valor VA', '0'))
    vr_mensal = _to_decimal(func_dados.get('vr_mensal', '0'))
    vt_diario = _to_decimal(func_dados.get('VT Diário', '0'))
        
    valor_dia = salario / Decimal("30")
    
    dt_fim_45 = dt_adm + timedelta(days=44)
    dt_fim_90 = dt_adm + timedelta(days=89)
    dias_trabalhados_hoje = (data_desligamento - dt_adm).days + 1
    
    if dias_trabalhados_hoje <= 45:
        fase = "1º Período de Experiência (45 dias)"
        dt_alvo = dt_fim_45
        dias_restantes = (dt_fim_45 - data_desligamento).days
        total_dias_projetados = 45
    elif dias_trabalhados_hoje <= 90:
        fase = "2º Período de Experiência (90 dias)"
        dt_alvo = dt_fim_90
        dias_restantes = (dt_fim_90 - data_desligamento).days
        total_dias_projetados = 90
    else:
        fase = "Contrato por Prazo Indeterminado"
        dt_alvo = None; dias_restantes = 0; total_dias_projetados = dias_trabalhados_hoje

    multa_479 = ((Decimal(str(dias_restantes)) * valor_dia) / Decimal("2")) if dias_restantes > 0 else Decimal("0.00")
    
    # 1. AVOS E FGTS (CENÁRIO A)
    meses_cheios_hoje = dias_trabalhados_hoje // 30
    dias_sobra_hoje = dias_trabalhados_hoje % 30
    avos_hoje = meses_cheios_hoje + (1 if dias_sobra_hoje >= 15 else 0)
    
    ferias_hoje = (salario / Decimal("12")) * Decimal(str(avos_hoje))
    terco_hoje = ferias_hoje / Decimal("3")
    decimo_hoje = (salario / Decimal("12")) * Decimal(str(avos_hoje))
    
    # Estimativa de Multa FGTS 40% (Saldo Base = Salários pagos + 13º proporcional)
    base_fgts_acumulada = (valor_dia * Decimal(str(dias_trabalhados_hoje))) + decimo_hoje
    saldo_fgts_est = base_fgts_acumulada * Decimal("0.08")
    multa_fgts_hoje = saldo_fgts_est * Decimal("0.40")
    
    # Custo de Benefícios Consumidos (Cenário A)
    dias_uteis_trabalhados = Decimal(str(int(dias_trabalhados_hoje * 0.73)))
    vt_gasto_hoje = dias_uteis_trabalhados * vt_diario
    va_gasto_hoje = (va_mensal / Decimal("30")) * Decimal(str(dias_trabalhados_hoje))
    vr_gasto_hoje = (vr_mensal / Decimal("30")) * Decimal(str(dias_trabalhados_hoje))

    # 2. AVOS PROJETADOS (CENÁRIO B)
    meses_cheios_proj = total_dias_projetados // 30
    dias_sobra_proj = total_dias_projetados % 30
    avos_proj = meses_cheios_proj + (1 if dias_sobra_proj >= 15 else 0)
    
    ferias_proj = (salario / Decimal("12")) * Decimal(str(avos_proj))
    terco_proj = ferias_proj / Decimal("3")
    decimo_proj = (salario / Decimal("12")) * Decimal(str(avos_proj))
    
    # 3. FATIAMENTO DE CAIXA (CENÁRIO B)
    salario_folha_mensal = Decimal("0.00")
    salario_rescisao_futura = Decimal("0.00")
    dias_rescisao_futura = 0
    
    if dias_restantes > 0:
        import calendar
        ultimo_dia_mes_atual = calendar.monthrange(data_desligamento.year, data_desligamento.month)[1]
        dias_ate_fim_mes = (date(data_desligamento.year, data_desligamento.month, ultimo_dia_mes_atual) - data_desligamento).days
        
        if dt_alvo.month == data_desligamento.month:
            salario_folha_mensal = Decimal(str(dias_restantes)) * valor_dia
            dias_rescisao_futura = 0
        else:
            salario_folha_mensal = Decimal(str(dias_ate_fim_mes)) * valor_dia
            dias_rescisao_futura = dias_restantes - dias_ate_fim_mes
            salario_rescisao_futura = Decimal(str(dias_rescisao_futura)) * valor_dia

    # VA/VR Futuros (Apenas para os dias do mês de rescisão)
    va_projetado = (va_mensal / Decimal("30")) * Decimal(str(dias_rescisao_futura))
    vr_projetado = (vr_mensal / Decimal("30")) * Decimal(str(dias_rescisao_futura))
    
    return {
        "dias_trabalhados_hoje": dias_trabalhados_hoje, "fase": fase,
        "dt_alvo": dt_alvo.strftime('%d/%m/%Y') if dt_alvo else "N/A",
        "dias_restantes": dias_restantes, "multa_479": multa_479.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP),
        
        "avos_hoje": avos_hoje, 
        "ferias_hoje": ferias_hoje.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP), 
        "terco_hoje": terco_hoje.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP), 
        "decimo_hoje": decimo_hoje.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP),
        "multa_fgts_hoje": multa_fgts_hoje.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP),
        "vt_gasto_hoje": vt_gasto_hoje.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP), 
        "va_gasto_hoje": va_gasto_hoje.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP), 
        "vr_gasto_hoje": vr_gasto_hoje.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP),
        
        "avos_proj": avos_proj, 
        "ferias_proj": ferias_proj.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP), 
        "terco_proj": terco_proj.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP), 
        "decimo_proj": decimo_proj.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP),
        
        "salario_folha_mensal": salario_folha_mensal.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP),
        "salario_rescisao_futura": salario_rescisao_futura.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP), 
        "dias_rescisao_futura": dias_rescisao_futura,
        "va_projetado": va_projetado.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP), 
        "vr_projetado": vr_projetado.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP),
        "va_mensal_cheio": va_mensal, "vr_mensal_cheio": vr_mensal,
        
        "salario_saldo_hoje": (Decimal(str(data_desligamento.day)) * valor_dia).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    }

# ==========================================
# MÓDULO: ORGANIZADOR DE SCANS
# ==========================================

def classificar_e_splitar_pdf(pdf_file):
    genai.configure(api_key=API_KEY_GEMINI)
    model = genai.GenerativeModel('gemini-2.5-flash', generation_config={"response_mime_type": "application/json"})
    
    with open("temp_scan.pdf", "wb") as f:
        f.write(pdf_file.getbuffer())
    
    arquivo_gemini = genai.upload_file(path="temp_scan.pdf")
    while arquivo_gemini.state.name == "PROCESSING":
        time.sleep(10)
        arquivo_gemini = genai.get_file(arquivo_gemini.name)

    prompt = """
    Analise este PDF de admissão e identifique todos os documentos presentes.
    
    REGRAS OBRIGATÓRIAS:
    1. Cada página do PDF pertence a APENAS UM documento. NUNCA coloque o mesmo número de página em dois documentos diferentes.
    2. SEPARE rigorosamente Certificados de Avaliações/Testes. O "Certificado NR 35" é um documento, e o "Teste/Avaliação NR 35" é outro documento totalmente diferente. Não agrupe.
    3. Retorne uma matriz exata de páginas (ex: [4, 5]).
    
    Para cada documento encontrado, retorne:
    1. O tipo de documento (ex: NR01 , NR06 , NR12 , NR18 ,NR35, ASO)
    2. O nome do funcionário.
    3. A lista de páginas.
    
    Retorne no formato JSON:
    [
      {""tipo": "TIPO DO DOCUMENTO", "funcionario": "NOME DO FUNCIONARIO", "paginas": [1, 2]}
    ]
    """
    response = model.generate_content([prompt, arquivo_gemini], request_options={"timeout": 600})
    mapa_documentos = json.loads(response.text)

    reader = PdfReader("temp_scan.pdf")
    zip_buffer = io.BytesIO()
    
    paginas_utilizadas = set() 
    
    with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as zip_file:
        for doc in mapa_documentos:
            writer = PdfWriter()
            
            paginas_deste_doc = sorted(list(set(doc['paginas'])))
            paginas_adicionadas = []
            
            for p in paginas_deste_doc:
                if 0 < p <= len(reader.pages) and p not in paginas_utilizadas:
                    writer.add_page(reader.pages[p-1])
                    paginas_utilizadas.add(p)
                    paginas_adicionadas.append(str(p))
            
            if paginas_adicionadas:
                output_pdf = io.BytesIO()
                writer.write(output_pdf)
                
                nome_func = doc.get("funcionario", "Desconhecido").replace(" ", "_").replace("/", "-")
                tipo_doc = doc.get("tipo", "Documento").replace(" ", "_").replace("/", "-")
                nome_arquivo = f"{nome_func}_{tipo_doc}.pdf"
                
                zip_file.writestr(nome_arquivo, output_pdf.getvalue())
                
    return zip_buffer.getvalue()

# ==========================================
# 6. FRONT-END
# ==========================================
st.sidebar.title("🏗️ Vicelos System")

# Diagnóstico rápido de conexão com Postgres
with st.sidebar.expander("🔌 Diagnóstico"):
    from db_client import health_check
    if st.button("Testar Postgres"):
        ok, msg = health_check()
        if ok:
            st.success("Conexão OK")
        else:
            st.error(f"Falhou: {msg}")
    st.caption("Use DB_HOST/PORT/NAME/USER/PASSWORD em env se necessário.")

menu = st.sidebar.radio("Navegue para:", ["👥 Admissão Inteligente", "💰 Folha de Pagamento", "📂 Organizador de Scans", "⚖️ Simulador de Desligamento", "🏦 Conciliação Bancária", "🧾 Gestão Fiscal (NFS-e)"])
st.sidebar.markdown("---")
st.sidebar.write("🗄️ Ferramentas")
abrir_sql = st.sidebar.checkbox("Abrir painel SQL rápido")

if abrir_sql:
    st.markdown("### 🗄️ Painel SQL (leitura)")
    st.caption("Roda SELECTs direto no Postgres local. Somente leitura.")
    default_query = "SELECT id, data_competencia, descricao, numero_documento FROM lancamentos ORDER BY criado_em DESC LIMIT 5;"
    sql_text = st.text_area("SQL", value=default_query, height=120)
    if st.button("Executar SQL"):
        try:
            import psycopg2
            import pandas as pd
            conn = psycopg2.connect(
                host=os.getenv("DB_HOST", "localhost"),
                port=os.getenv("DB_PORT", "5432"),
                dbname=os.getenv("DB_NAME", "vicelos_erp"),
                user=os.getenv("DB_USER", "renancarvalho"),
                password=os.getenv("DB_PASSWORD", ""),
            )
            df = pd.read_sql(sql_text, conn)
            st.dataframe(df)
        except Exception as e:
            st.error(f"Erro ao executar SQL: {e}")

if menu == "👥 Admissão Inteligente":
    st.title("🏗️ Vicelos - Admissão Inteligente")
    if st.session_state.ultimo_kit:
        st.success("✅ Processo Finalizado com Sucesso!"); st.markdown(f"**Pasta:** [Clique para Abrir]({st.session_state.ultimo_kit})")
        if st.button("🔄 Iniciar Nova Admissão"): st.session_state.ultimo_kit = None; st.session_state.dados_extraidos = None; st.rerun()
    else:
        uploaded_files = st.file_uploader("Documentos", accept_multiple_files=True, type=['pdf', 'png', 'jpg', 'jpeg'])
        if uploaded_files and st.button("🔍 Analisar Documentos"):
            with st.spinner('Lendo...'): st.session_state.dados_extraidos = processar_documentos_ia(uploaded_files); st.success("Pronto!")
        
        if st.session_state.dados_extraidos:
            d = st.session_state.dados_extraidos
            with st.form("form_adm"):
                st.subheader("📝 Dados Pessoais")
                c_n1, c_n2 = st.columns([3, 1])
                nome = c_n1.text_input("Nome Completo", d.get("Nome Completo", ""))
                cpf = c_n2.text_input("CPF", d.get("CPF", ""))
                
                c_r1, c_r2, c_r3, c_r4 = st.columns(4)
                rg = c_r1.text_input("RG / CRNM", d.get("RG", ""))
                pis = c_r2.text_input("PIS *", d.get("PIS", ""))
                nacionalidade = c_r3.text_input("Nacionalidade", d.get("Nacionalidade", "Brasileiro"))
                est_civil = c_r4.selectbox("Estado Civil", ["Solteiro(a)", "Casado(a)", "Divorciado(a)", "Viúvo(a)"])

                st.markdown("---")
                st.subheader("🏠 Endereço")
                c_e1, c_e2, c_e3 = st.columns([1, 2, 1])
                cep = c_e1.text_input("CEP", d.get("CEP", ""))
                logradouro = c_e2.text_input("Logradouro", d.get("Logradouro", ""))
                numero = c_e3.text_input("Número", d.get("Numero Endereco", ""))

                c_e4, c_e5, c_e6, c_e7 = st.columns([1, 1, 1, 1])
                complemento = c_e4.text_input("Complemento", d.get("Complemento", ""))
                bairro = c_e5.text_input("Bairro", d.get("Bairro", ""))
                cidade = c_e6.text_input("Cidade", d.get("Cidade", ""))
                estado = c_e7.text_input("UF", d.get("Estado", ""))

                st.markdown("---")
                st.subheader("🏗️ Dados do Contrato")
                c_c1, c_c2, c_c3 = st.columns(3)
                cargo = c_c1.text_input("Cargo", "Pintor de Obras")
                cbo = c_c2.text_input("CBO", "7166-10")
                obra_selecionada = c_c3.selectbox("Obra", ["LARIS BURITIS", "Residencial Alphaville", "Galpão Logístico SP"])
                
                c_c4, c_c5, c_c6 = st.columns(3)
                salario = c_c4.text_input("Salário (R$)", "2.664,75")
                vt_diario = c_c5.text_input("VT Diário (Ida+Volta) R$", "10,00")
                inicio = c_c6.text_input("Data Início", datetime.now().strftime('%d/%m/%Y'))
                
                if st.form_submit_button("🚀 Gerar Kit Admissional"):
                    dados_finais = {
                        'nome': nome, 'cpf': cpf, 'rg': rg, 'pis': pis, 
                        'cargo': cargo, 'cbo': cbo, 'salario': salario, 
                        'VT Diário': vt_diario,
                        'data_inicio': inicio, 'obra': obra_selecionada, 
                        'nacionalidade': nacionalidade, 'estado_civil': est_civil,
                        'Logradouro': logradouro, 'Numero Endereco': numero, 
                        'Bairro': bairro, 'Cidade': cidade, 'Estado': estado, 
                        'CEP': cep, 'Complemento': complemento,
                        **d 
                    }
                    with st.spinner("Gerando..."): 
                        st.session_state.ultimo_kit = gerar_kit_admissional(dados_finais)
                        st.rerun()

elif menu == "💰 Folha de Pagamento":
    st.title("💰 Folha de Pagamento")
    aba1, aba2 = st.tabs(["👤 Individual", "🏭 Em Lote"])
    with aba1:
        if st.session_state.ultimo_holerite:
            st.success("✅ Gerado!"); st.markdown(f"**Doc:** [Abrir]({st.session_state.ultimo_holerite})")
            if st.button("🔄 Novo"): st.session_state.ultimo_holerite = None; st.rerun()
        else:
            if st.button("🔄 Carregar Lista"): st.session_state.funcs = buscar_funcionarios()
            if st.session_state.funcs:
                opcoes = [f"{f['nome']} | {f['cargo']}" for f in st.session_state.funcs]
                func = st.session_state.funcs[opcoes.index(st.selectbox("Funcionário", opcoes))]
                with st.form("f_ind"):
                    col_t, col_f = st.columns(2)
                    tipo_p = col_t.radio("Tipo", ["Fechamento Mensal", "Adiantamento Quinzenal (Dia 20)"])
                    data_f = col_f.date_input("Fechamento", value=date(2026, 2, 28), format="DD/MM/YYYY")
                    c1, c2, c3 = st.columns(3)
                    he = c1.number_input("HE 60%", 0.0); he100 = c2.number_input("HE 100%", 0.0); atraso = c3.number_input("Atrasos", 0.0)
                    
                    c4, c5, c6, c7, c8 = st.columns(5)
                    falta = c4.number_input("Faltas", 0)
                    dsr = c5.number_input("DSR Perdido", 0)
                    dias_vt = c6.number_input("Dias Úteis VT", value=22, min_value=0, max_value=31)
                    
                    # Se for adiantamento, não faz sentido marcar "Pagar Adiantamento" (ele JÁ É o adiantamento)
                    p_vale = c7.checkbox("Pagar Adiant.?", value=True, disabled=(tipo_p == "Adiantamento Quinzenal (Dia 20)"))
                    d_cesta = c8.checkbox("Desc. VA?", value=True)
                    
                    if st.form_submit_button("Gerar"):
                        with st.spinner("Gerando..."):
                            link, _, _ = gerar_holerite_dinamico({
                                **func, 
                                'salario_base': func['salario'], 
                                'data_fechamento': data_f.strftime('%d/%m/%Y'), 
                                'qtd_he': he, 
                                'qtd_he_100': he100, 
                                'dias_faltas': falta, 
                                'dsr_descontado': dsr, 
                                'horas_atrasos': atraso, 
                                'dias_uteis_vt': dias_vt,
                                'Desc. Adiant.?': p_vale, 
                                'Desc. VA?': d_cesta, 
                                'tipo_processamento': tipo_p
                            })
                            st.session_state.ultimo_holerite = link; st.rerun()
    with aba2:
        c_t, c_d, c_p, c_f = st.columns([2, 1, 1, 1])
        t_lote = c_t.radio("Lote:", ["Fechamento Mensal", "Adiantamento Quinzenal (Dia 20)"], horizontal=True, key="tipo_lote_radio")
        d_deb = c_d.date_input("Débito", value=date.today(), format="DD/MM/YYYY", key="data_debito_lote")
        d_pag = c_p.date_input("Crédito", value=date.today(), format="DD/MM/YYYY", key="data_pagto_lote")
        
        # Calcula o último dia do mês corrente como padrão sensível
        hoje = date.today()
        proximo_mes = hoje.replace(day=28) + timedelta(days=4)
        ultimo_dia_mes = proximo_mes - timedelta(days=proximo_mes.day)
        data_fechamento_lote = c_f.date_input("Fechamento", value=ultimo_dia_mes, format="DD/MM/YYYY", key="data_fechamento_lote_widget")
        
        if st.button("📊 Carregar Tabela"):
             df = pd.DataFrame(buscar_funcionarios())
             if not df.empty:
                 df.insert(0, 'Selecionar', True)
                 # Garante que a data de fechamento segue o input do usuário na hora de carregar
                 df['Data Fechamento'] = data_fechamento_lote.strftime('%d/%m/%Y')
                 df['HE 60%'] = 0.0; df['HE 100%'] = 0.0; df['Faltas'] = 0; df['DSR Perdido'] = 0; df['Atrasos (Hrs)'] = 0.0; 
                 
                 # Cálculo Automático de Dias Úteis VT para novos admitidos
                 df['Dias Úteis VT'] = df['admissao'].apply(lambda x: calcular_dias_uteis_proporcionais(x, data_fechamento_lote))
                 if 'VT Diário' not in df.columns: df['VT Diário'] = 0.0
                 df['Desc. Adiant.?'] = False if t_lote == "Adiantamento Quinzenal (Dia 20)" else True
                 if 'Descontar VT' not in df.columns: df['Descontar VT'] = True
                 if 'Descontar VA' not in df.columns: df['Descontar VA'] = True

                 # Reordenar colunas para melhor visualização (grupos de descontos lado a lado)
                 ordem_visual = [
                     'Selecionar', 'nome', 'cpf', 'admissao', 'salario', 
                     'VT Diário', 'Descontar VT', 'Dias Úteis VT', 
                     'Valor VA', 'Descontar VA', 'Valor VR', 
                     'Desc. Adiant.?', 'Data Fechamento', 
                     'HE 60%', 'HE 100%', 'Faltas', 'DSR Perdido', 'Atrasos (Hrs)',
                     'banco', 'agencia', 'conta', 'obra'
                 ]
                 cols_final = [c for c in ordem_visual if c in df.columns]
                 cols_resto = [c for c in df.columns if c not in cols_final]
                 df = df[cols_final + cols_resto]

             st.session_state.df_lote = df
        if st.session_state.df_lote is not None and not st.session_state.df_lote.empty:
            df_ed = st.data_editor(
                st.session_state.df_lote, 
                hide_index=True,
                column_config={
                    "banco": st.column_config.TextColumn("Banco"),
                    "agencia": st.column_config.TextColumn("Agência"),
                    "conta": st.column_config.TextColumn("Conta")
                }
            )
            if st.button("🚀 Processar Geral"):
                links = []; l_inter = []; t_folha = Decimal("0.00"); bar = st.progress(0)
                df_selecionados = df_ed[df_ed['Selecionar'] == True]
                if df_selecionados.empty:
                    st.warning("Nenhum funcionário selecionado.")
                else:
# --- NOVO: COLETA DE LOGS PARA O SHEETS ---
                    logs_financeiros = []
# -----------------------------------------
                    for i, row in df_selecionados.iterrows():
                        link, liq, det = gerar_holerite_dinamico({
                            **row, 
                            'salario_base': row['salario'], 
                            'data_fechamento': row['Data Fechamento'], 
                            'qtd_he': row['HE 60%'], 
                            'qtd_he_100': row['HE 100%'], 
                            'dias_faltas': row['Faltas'], 
                            'dsr_descontado': row['DSR Perdido'], 
                            'horas_atrasos': row['Atrasos (Hrs)'], 
                            'Dias Úteis VT': row['Dias Úteis VT'],
                            'VT Diário': row.get('VT Diário', 0),
                            'Descontar VT': row.get('Descontar VT', True),
                            'Desc. Adiant.?': row['Desc. Adiant.?'], 
                            'Descontar VA': row['Descontar VA'], 
                            'Valor VA': row.get('Valor VA', 0),
                            'tipo_processamento': t_lote
                        })
                        t_folha += liq; l_inter.append({"Nome": row['nome'], "CPF": row['cpf'], "Conta": row['conta'], "Valor": round(float(liq), 2), "Data Debito": d_deb.strftime('%d/%m/%Y'), "Data Pagamento": d_pag.strftime('%d/%m/%Y')})
                        links.append(f"- **{row['nome']}:** [Holerite]({link})"); bar.progress((i+1)/len(df_selecionados))
                        
                        # --- INTEGRAÇÃO ERP: PERSISTÊNCIA NO BANCO DE DADOS ---
                        try:
                            # Se tivermos o entidade_id (vindo do Postgres), salvamos diretamente
                            ent_id = row.get('entidade_id')
                            if ent_id:
                                save_payroll_record(
                                    entidade_id=ent_id,
                                    competencia=row['Data Fechamento'],
                                    data_pagamento=d_pag,
                                    valor_liquido=liq,
                                    obra=row.get('obra', '')
                                )
                        except Exception as e:
                            st.error(f"Erro ao persistir lançamento no banco: {e}")
                        # -----------------------------------------------------
                        
                        # --- COLETAR LOG DETALHADO PARA O SHEETS ---
                        logs_financeiros.append({
                            "competencia": row['Data Fechamento'],
                            "cpf": row['cpf'],
                            "nome": row['nome'],
                            "salario_base": det['salario_base'],
                            "he_total": det['he_60'] + det['he_100'] + det['dsr'],
                            "bruto": det['bruto'],
                            "inss": det['inss'],
                            "irrf": det['irrf'],
                            "vt": det['vt'],
                            "va": det['va'],
                            "faltas": det['faltas'],
                            "atrasos": det['atrasos'],
                            "sindical": det['sindical'],
                            "adiantamento": det['adiantamento'],
                            "desconto_total": det['desc_total'],
                            "liquido": det['liquido'],
                            "fgts": det['fgts'],
                            "banco": "Postgres" if row.get('entidade_id') else "Google Sheets"
                        })
                        # ---------------------------------
                        
                        # --- INTEGRAÇÃO ERP: POST DO HOLERITE (Legado/API) ---
                        import requests
                        try:
                            payload_folha = {
                                "cpf_funcionario": str(row['cpf']).replace(".", "").replace("-", "").strip(),
                                "competencia_texto": row['Data Fechamento'],
                                "data_pagamento": d_pag.strftime('%Y-%m-%d'),
                                "salario_liquido": float(liq)
                            }
                            # Envia para a porta 8000 (onde o Uvicorn estará rodando em background)
                            requests.post("http://localhost:8000/lancamentos/folha", json=payload_folha, timeout=3)
                        except:
                            pass # Em produção deveríamos logar.
                        # ----------------------------------------
                    
                    # --- NOVO: SALVAR LOG NO GOOGLE SHEETS ---
                    if logs_financeiros:
                        with st.spinner("Salvando log no Google Sheets..."):
                            salvar_log_folha_sheets(logs_financeiros)
                    # ------------------------------------------

                st.download_button("📥 Baixar Planilha Inter", gerar_excel_banco_inter(l_inter, float(t_folha)), f"Folha_{datetime.now().strftime('%Y%m%d')}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                st.markdown("\n".join(links))

elif menu == "📂 Organizador de Scans":
    st.title("📂 Organizador Inteligente de Documentos")
    st.info("Jogue aqui o PDF único com todas as páginas escaneadas. O sistema vai separar, renomear e entregar um arquivo .ZIP pronto para envio.")
    
    arquivo_scan = st.file_uploader("Upload do Scan Completo", type=['pdf'])
    
    if arquivo_scan and st.button("🪄 Desmembrar e Renomear"):
        with st.spinner("A IA está analisando as páginas e identificando os documentos..."):
            zip_data = classificar_e_splitar_pdf(arquivo_scan)
            
            st.success("✅ Tudo pronto! Arquivos separados com sucesso.")
            st.download_button(
                label="📥 Baixar Documentos Separados (.ZIP)",
                data=zip_data,
                file_name=f"Documentos_Separados.zip",
                mime="application/zip"
            )

elif menu == "⚖️ Simulador de Desligamento":
    st.title("⚖️ Simulador de Cenários de Desligamento")
    st.info("Analise os custos operacionais e a linha do tempo legal antes de aprovar a rescisão com a diretoria.")
    
    if st.button("🔄 Puxar Dados da Planilha"):
        st.session_state.funcs = buscar_funcionarios()
    
    if st.session_state.funcs:
        opcoes = [f"{f['nome']} | Salário: R$ {f['salario']}" for f in st.session_state.funcs]
        selecao = st.selectbox("1. Selecione o Colaborador", opcoes)
        func_selecionado = st.session_state.funcs[opcoes.index(selecao)]
        
        dt_proj_desligamento = st.date_input("2. Data Projetada para o Desligamento", datetime.now().date())
        
        resultados = calcular_cenarios_desligamento(func_selecionado, dt_proj_desligamento)
        
        if resultados:
            st.markdown("---")
            st.subheader("📊 Diagnóstico Contratual")
            col1, col2, col3 = st.columns(3)
            col1.metric("Fase do Contrato", resultados["fase"])
            
            # CORREÇÃO: A chave agora se chama 'dias_trabalhados_hoje'
            col2.metric("Dias Trabalhados", resultados["dias_trabalhados_hoje"]) 
            
            if resultados["dias_restantes"] > 0:
                col3.metric("Dias Faltantes p/ Término", resultados["dias_restantes"], delta=f"Término ideal em {resultados['dt_alvo']}")
            else:
                col3.metric("Status", "Fora da Experiência")
            
            st.markdown("---")
            st.subheader("⚖️ Comparativo Financeiro (Tomada de Decisão)")
            
            c_hoje, c_manter = st.columns(2)
            
            with c_hoje:
                st.warning(f"**CENÁRIO A: Desligar Hoje ({dt_proj_desligamento.strftime('%d/%m/%Y')})**")
                
                st.markdown("**Verbas Rescisórias a Pagar:**")
                st.markdown(f"&nbsp;&nbsp;&nbsp;↳ Multa Art. 479 (50%): R$ {formatar_moeda(resultados['multa_479'])}")
                st.markdown(f"&nbsp;&nbsp;&nbsp;↳ Multa 40% FGTS (Est.): R$ {formatar_moeda(resultados['multa_fgts_hoje'])}")
                st.markdown(f"&nbsp;&nbsp;&nbsp;↳ Saldo Salário ({dt_proj_desligamento.day} dias): R$ {formatar_moeda(resultados['salario_saldo_hoje'])}")
                
                if resultados['avos_hoje'] > 0:
                    st.markdown(f"&nbsp;&nbsp;&nbsp;↳ Férias ({resultados['avos_hoje']}/12): R$ {formatar_moeda(resultados['ferias_hoje'])}")
                    st.markdown(f"&nbsp;&nbsp;&nbsp;↳ 1/3 Férias: R$ {formatar_moeda(resultados['terco_hoje'])}")
                    st.markdown(f"&nbsp;&nbsp;&nbsp;↳ 13º Salário ({resultados['avos_hoje']}/12): R$ {formatar_moeda(resultados['decimo_hoje'])}")
                else:
                    st.markdown(f"&nbsp;&nbsp;&nbsp;↳ Férias e 13º: R$ 0,00 *(Não atingiu 15 dias)*")
                
                st.markdown("**Custo de Benefícios Consumidos (Trabalhados):**")
                st.markdown(f"&nbsp;&nbsp;&nbsp;↳ VT Utilizado: R$ {formatar_moeda(resultados['vt_gasto_hoje'])}")
                st.markdown(f"&nbsp;&nbsp;&nbsp;↳ VA Utilizado: R$ {formatar_moeda(resultados['va_gasto_hoje'])}")
                st.markdown(f"&nbsp;&nbsp;&nbsp;↳ VR Utilizado: R$ {formatar_moeda(resultados['vr_gasto_hoje'])}")
                
                custo_total_a = (
                    resultados['multa_479'] + resultados['multa_fgts_hoje'] + 
                    resultados['salario_saldo_hoje'] + resultados['ferias_hoje'] + 
                    resultados['terco_hoje'] + resultados['decimo_hoje'] + 
                    resultados['vt_gasto_hoje'] + resultados['va_gasto_hoje'] + 
                    resultados['vr_gasto_hoje']
                )
                st.markdown(f"### Custo Total Histórico + Rescisão: R$ {formatar_moeda(custo_total_a)}")

            with c_manter:
                if resultados["dias_restantes"] > 0:
                    st.success(f"**CENÁRIO B: Manter até ({resultados['dt_alvo']})**")
                    st.markdown(f"- **Multa Art. 479 (50%):** R$ 0,00 (ISENTO)")
                    
                    st.markdown("**1. Rescisão Futura:**")
                    st.markdown(f"&nbsp;&nbsp;&nbsp;↳ Férias ({resultados['avos_proj']}/12): R$ {formatar_moeda(resultados['ferias_proj'])}")
                    st.markdown(f"&nbsp;&nbsp;&nbsp;↳ 1/3 Férias: R$ {formatar_moeda(resultados['terco_proj'])}")
                    st.markdown(f"&nbsp;&nbsp;&nbsp;↳ 13º Salário ({resultados['avos_proj']}/12): R$ {formatar_moeda(resultados['decimo_proj'])}")
                    st.markdown(f"&nbsp;&nbsp;&nbsp;↳ Saldo Salário mês seguinte ({resultados['dias_rescisao_futura']} dias): R$ {formatar_moeda(resultados['salario_rescisao_futura'])}")
                    
                    subtotal_rescisao = resultados['ferias_proj'] + resultados['terco_proj'] + resultados['decimo_proj'] + resultados['salario_rescisao_futura']
                    st.markdown(f"**&nbsp;&nbsp;&nbsp;= Total Rescisão Futura: R$ {formatar_moeda(subtotal_rescisao)}**")
                    
                    st.markdown("**2. Fluxo de Folha Normal (Até 5º Dia Útil):**")
                    st.markdown(f"&nbsp;&nbsp;&nbsp;↳ Salário do mês atual: R$ {formatar_moeda(resultados['salario_folha_mensal'])}")
                    
                    st.markdown("**3. Custo de VA/VR (Apenas mês seguinte):**")
                    if resultados['dias_rescisao_futura'] > 0:
                        st.markdown(f"&nbsp;&nbsp;&nbsp;↳ Vale Alimentação ({resultados['dias_rescisao_futura']} dias): R$ {formatar_moeda(resultados['va_projetado'])}")
                        st.markdown(f"&nbsp;&nbsp;&nbsp;↳ Vale Refeição ({resultados['dias_rescisao_futura']} dias): R$ {formatar_moeda(resultados['vr_projetado'])}")
                    else:
                        st.markdown("&nbsp;&nbsp;&nbsp;↳ *Sem novos desembolsos (término ocorre no mês atual)*")
                    
                    custo_total_b = subtotal_rescisao + resultados['salario_folha_mensal'] + resultados['va_projetado'] + resultados['vr_projetado']
                    st.markdown(f"### Desembolso Adicional Total: R$ {formatar_moeda(custo_total_b)}")
                else:
                    st.info("Colaborador fora do período de experiência.")
            
            # --- NOVO BLOCO: BALANÇO FINANCEIRO ---
            if resultados["dias_restantes"] > 0:
                st.markdown("<br>", unsafe_allow_html=True)
                diferenca = custo_total_b - custo_total_a
                
                st.info(f"**💡 Balanço Financeiro (Custo Total B - Custo Total A):** A diferença de desembolso entre os dois cenários é de **R$ {formatar_moeda(diferenca)}**\n\n*Nota: O Cenário B não leva em consideração o valor gerado pela produção do funcionário no período em que ele permanecer com o contrato ativo*")
            # --------------------------------------

            st.markdown("---")
            st.subheader("🚨 Prazos Críticos de Conformidade")
            prazo_pagamento = dt_proj_desligamento + timedelta(days=10)
            
            st.error(f"**1. Pagamento PIX:** As verbas rescisórias devem cair na conta do funcionário até **{prazo_pagamento.strftime('%d/%m/%Y')}** (10 dias corridos).")
            st.error(f"**2. e-Social:** O evento S-2299 deve ser enviado no portal web em até 10 dias do desligamento (geralmente junto com o pagamento) para a emissão oficial do TRCT.")

elif menu == "🏦 Conciliação Bancária":
    st.title("🏦 Conciliação Inteligente (Banco Inter)")
    st.info("Central do Robô Financeiro: Faça o upload do Extrato .OFX ou .CSV. O algoritmo irá varrer suas Contas a Pagar/Receber pendentes e efetivar a baixa automática no fluxo de caixa real da empresa.")
    
    arquivo_upload = st.file_uploader("Arraste e solte o Extrato aqui", type=["ofx", "csv", "xlsx"])
    
    if arquivo_upload:
        st.write(f"📄 Arquivo pronto: `{arquivo_upload.name}`")
        st.warning("Certifique-se que o motor do ERP (FastAPI) está ligado no terminal antes de processar.")
        if st.button("🚀 Iniciar Motor de Conciliação"):
            with st.spinner("Descriptografando Extrato e caçando Títulos Pendentes..."):
                try:
                    import requests
                    files = {"file": (arquivo_upload.name, arquivo_upload.getvalue())}
                    resultado = requests.post("http://localhost:8000/integracoes/inter/conciliar", files=files)
                    
                    if resultado.status_code == 200:
                        dados = resultado.json()
                        st.success(f"✅ {dados.get('message', 'Concluído!')}")
                        
                        col1, col2, col3 = st.columns(3)
                        col1.metric("Linhas do Extrato", dados['analitico']['linhas_processadas'])
                        col2.metric("Matches (Baixas)", f"✅ {dados['analitico']['baixas_automaticas']}")
                        col3.metric("Não Conciliados", f"⚠️ {dados['analitico']['nao_conciliados']}")
                        
                        st.json(dados)
                    else:
                        st.error(f"❌ O Backend Recusou o Arquivo: {resultado.text}")
                except Exception as e:
                    st.error(f"❌ Erro Crítico de Comunicação. O terminal do FastAPI está ligado? Detalhe: {e}")

    st.markdown("---")
    st.markdown("#### 💳 Processamento de Faturas de Cartão")
    st.info("Caso queira lançar as dívidas do Cartão de Crédito como Despesas na contabilidade, use o CSV da Fatura abaixo:")
    
    fatura_upload = st.file_uploader("Upload CSV Fatura", type=["csv"], key="fatura_upload")
    if fatura_upload:
        if st.button("💸 Lançar Despesas do Cartão"):
            with st.spinner("Fragmentando Compras e criando Dívida Analítica no Contas a Pagar..."):
                try:
                    import requests
                    files = {"file": (fatura_upload.name, fatura_upload.getvalue())}
                    resultado = requests.post("http://localhost:8000/integracoes/inter/fatura", files=files)
                    
                    if resultado.status_code == 201:
                        dados = resultado.json()
                        st.success("✅ Fatura importada com Sucesso para o Motor de Partidas Dobradas!")
                        st.metric("Total da Dívida a Pagar", f"R$ {dados['total_registrado']:.2f}")
                        st.metric("Lançamentos de Despesa Unificados", dados['despesas_criadas'])
                    else:
                        st.error(f"❌ Erro do Servidor: {resultado.text}")
                except Exception as e:
                    st.error(f"❌ Erro Crítico do Back-end: {e}")

    st.markdown("---")
    st.subheader("📋 Títulos Pendentes (Não Conciliados)")
    if st.button("🔄 Atualizar Lista de Pendências"):
        try:
            r = requests.get("http://localhost:8000/financeiro/pendencias")
            if r.status_code == 200:
                pendencias = r.json()
                if pendencias:
                    df_p = pd.DataFrame(pendencias)
                    st.table(df_p[["tipo", "valor", "data_vencimento", "entidade", "descricao"]])
                else:
                    st.success("Tudo em dia! Nenhuma pendência encontrada.")
        except:
            st.error("Não foi possível conectar ao Backend para buscar pendências.")

elif menu == "🧾 Gestão Fiscal (NFS-e)":
    st.title("🧾 Gestão Fiscal: Importação de NFS-e")
    st.info("Importação local de NFS-e (CSV prefeitura). O arquivo será lido e os lançamentos serão gravados direto no Postgres (R01/A01/A02/A03) com movimentação de entrada opcional.")
    
    csv_upload = st.file_uploader("Upload CSV NFS-e", type=["csv"])
    if csv_upload and st.button("📡 Importar no ERP"):
        with st.spinner("Processando CSV..."):
            import pandas as pd
            from decimal import Decimal
            from db_client import import_nfse_rows
            import re

            def parse_money(txt):
                if pd.isna(txt):
                    return Decimal("0")
                s = str(txt).replace(".", "").replace(",", ".")
                try:
                    return Decimal(s)
                except Exception:
                    return Decimal("0")

            def parse_date(txt):
                if pd.isna(txt) or not str(txt).strip():
                    return None
                for fmt in ("%d/%m/%Y", "%d/%m/%Y %H:%M:%S"):
                    try:
                        return datetime.strptime(str(txt).split()[0], fmt).date()
                    except Exception:
                        continue
                return None

            def extract_vencimento(desc: str):
                if not desc:
                    return None
                m = re.search(r"VENCIMENTO\\s*(\\d{2}/\\d{2}/\\d{4})", desc.upper())
                if m:
                    return parse_date(m.group(1))
                return None

            def extract_valor_liquido(desc: str):
                if not desc:
                    return Decimal("0")
                m = re.search(r"VALOR\\s*LIQUIDO\\s*\\.\\.\\.\\.\\.\\.\\.\\.\\.\\.\\s*R\\$\\s*([\\d.,]+)", desc.upper())
                if m:
                    return parse_money(m.group(1))
                m2 = re.search(r"VALOR\\s*LIQUIDO[^\\d]*([\\d.,]+)", desc.upper())
                if m2:
                    return parse_money(m2.group(1))
                return Decimal("0")

            def extract_ret_tec(desc: str):
                if not desc:
                    return Decimal("0")
                # priorizar valor monetário após R$
                m = re.search(r"RETEN[ÇC][AÃ]O\\s*TECNICA[^R$]*R\\$\\s*([\\d.,]+)", desc.upper())
                if m:
                    return parse_money(m.group(1))
                m2 = re.search(r"RETEN[ÇC][AÃ]O\\s*TECNICA[^\\d]*([\\d.,]+)", desc.upper())
                if m2:
                    return parse_money(m2.group(1))
                return Decimal("0")

            df = pd.read_csv(csv_upload, encoding="latin-1", sep=";")
            if "Tipo de Registro" in df.columns:
                df = df[df["Tipo de Registro"] == "2"]

            rows = []
            for _, r in df.iterrows():
                descricao_raw = str(r.get("Discriminação dos Serviços", "")).strip()
                vencimento = extract_vencimento(descricao_raw)

                # Valor a receber preferencial
                valor_a_receber = parse_money(r.get("Valor a Receber"))
                valor_total_recebido = parse_money(r.get("Valor Total Recebido"))
                valor_liquido_desc = extract_valor_liquido(descricao_raw)
                ret_tec = extract_ret_tec(descricao_raw)

                row = {
                    "numero": str(r.get("Nº NFS-e", "")).strip(),
                    "data_fato": parse_date(r.get("Data do Fato Gerador") or r.get("Data Hora NFE")),
                    "descricao": descricao_raw,
                    "tomador_nome": str(r.get("Razão Social do Tomador", "")).strip(),
                    "tomador_cpf": str(r.get("CPF/CNPJ do Tomador", "")).replace(".", "").replace("-", "").replace("/", "").strip(),
                    "valor_servicos": parse_money(r.get("Valor dos Serviços")),
                    "deducoes": parse_money(r.get("Valor das Deduções")),
                    "iss_retido": parse_money(r.get("ISS Retido")),
                    "inss_retido": parse_money(r.get("INSS")),
                    "valor_recebido": next(v for v in [valor_a_receber, valor_total_recebido, valor_liquido_desc] if v > 0) if any(v > 0 for v in [valor_a_receber, valor_total_recebido, valor_liquido_desc]) else Decimal("0"),
                    "retencao_tecnica": ret_tec,
                    "obra_codigo": str(r.get("Matrícula da Obra", "")).strip() or None,
                    "data_pagamento": vencimento or parse_date(r.get("Data do Fato Gerador") or r.get("Data Hora NFE")),
                }
                rows.append(row)

            if not rows:
                st.warning("Nenhuma linha de nota encontrada (Tipo de Registro != 2).")
            else:
                resumo = import_nfse_rows(rows)
                st.success(f"Lançamentos criados: {resumo['lancamentos_criados']} | Movimentações: {resumo['movimentacoes_criadas']}")
                if resumo.get("erros"):
                    st.warning("Ocorreram avisos/erros:")
                    for e in resumo["erros"]:
                        st.write(f"- {e}")
